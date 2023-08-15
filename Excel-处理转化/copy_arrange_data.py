from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd


def copy_arrange_data():

    # 读取 excel 文件
    excel_file = '/Users/xinwang/Desktop/zhengsiyu/get_log_data/jk埋点.xlsx'

    wb = load_workbook(excel_file)

    # 创建一个名为“格式化后埋点用例”的新工作表
    new_sheet = wb.create_sheet(title='格式化后埋点用例')

    # 获取所有工作表的名称
    sheet_names = wb.sheetnames

    # 选择第2个工作表
    sheet2 = wb[sheet_names[1]]

    # 提取第一列数据
    first_column = [cell.value for cell in sheet2['A']]

    # 获取第二列及其之后的数据
    remain_columns = []

    for column in sheet2.iter_cols(min_row=1, min_col=2, values_only=True):
        remain_columns.append(column)

    format_tuple = []

    for column in remain_columns:
        format_tuple.append(first_column)

        format_tuple.append(column)

    #  传入一个 工作表实例对象，和列数据格式：[1,2,3]， 给工作表中增加一个列

    def add_column_to_sheet(sheet, column_data, column_index):
        # 在每一行的第 max_column + 1 列插入对应的数据
        for row_num, data in enumerate(column_data, start=1):
            sheet.cell(row=row_num, column=column_index, value=data)

    for column_index, column_tuple in enumerate(format_tuple, start=1):
        add_column_to_sheet(new_sheet, column_tuple, column_index)

    # 保存工作簿，包含整理后的数据
    output_file = '/Users/xinwang/Desktop/zhengsiyu/get_log_data/jk埋点.xlsx'

    wb.save(output_file)


copy_arrange_data()
