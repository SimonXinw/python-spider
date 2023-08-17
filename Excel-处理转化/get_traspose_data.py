from openpyxl import load_workbook
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from config import SOURCE_DIR_PATH, SAVE_DIR_PATH, NEW_SHEET_TITLE
import pandas as pd
from utils import common


def get_traspose_data():

    utiles_instance = common.Utils()

    file_name = utiles_instance.get_first_filename(SOURCE_DIR_PATH)

    # excel 文件绝对路径 path
    source_file_path = os.path.abspath(
        os.path.join(SOURCE_DIR_PATH, file_name))

    # excel 文件绝对路径 path
    save_file_path = os.path.abspath(
        os.path.join(SAVE_DIR_PATH, file_name))

    wb = load_workbook(source_file_path)

    # 创建一个名为“埋点用例”的新工作表
    new_sheet = wb.create_sheet(title=NEW_SHEET_TITLE)

    # 获取所有工作表的名称
    sheet_names = wb.sheetnames

    # 选择第1个工作表
    selected_sheet = wb[sheet_names[0]]

    # 获取数据的起始行和结束行
    start_row = selected_sheet.min_row
    end_row = selected_sheet.max_row

    # 获取数据的起始列和结束列
    start_column = selected_sheet.min_column
    end_column = selected_sheet.max_column

    # 选择指定范围内的数据（从1开始）
    # minx_row：起始行数
    # max_row：结束行数
    # min_col：起始列数
    # max_col：结束列数

    # 第一行
    first_row_data = []

    for row in selected_sheet.iter_rows(min_row=start_row, max_row=start_row, values_only=True):
        first_row_data = row

    # 剩余行
    remain_rows_list = []

    for row in selected_sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_column,
                                        max_col=end_column, values_only=True):
        remain_rows_list.append(row)

    #  格式化全部行
    all_rows_list = []

    for column in remain_rows_list:
        all_rows_list.append(first_row_data)

        all_rows_list.append(column)

    # 将数据转换为DataFrame并进行转置
    df = pd.DataFrame(all_rows_list)

    transposed_df = df.transpose()

    # 将转置后的数据粘贴到新的工作表中
    for row in dataframe_to_rows(transposed_df, index=False, header=False):
        new_sheet.append(row)

    # 保存工作簿
    wb.save(save_file_path)

    return


get_traspose_data()
